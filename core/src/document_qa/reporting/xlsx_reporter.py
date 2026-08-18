"""验收报告导出为 XLSX 问题清单。

面向验收签核场景：Sheet1 逐条问题（含人工判定预留列），Sheet2 文档
摘要。列结构与 Issue 字段一一对应，判定列供复核数据回填。
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from document_qa.schemas import QAReport

# 严重度 → 单元格底色（浅色调，打印友好）。
_SEVERITY_FILL = {
    "critical": PatternFill("solid", fgColor="F4B6B0"),
    "high": PatternFill("solid", fgColor="F8D3CE"),
    "medium": PatternFill("solid", fgColor="FDF0D5"),
    "low": PatternFill("solid", fgColor="E3EFE7"),
    "info": PatternFill("solid", fgColor="ECEFF1"),
}


def export_xlsx(report: QAReport, output_path: Path) -> Path:
    """把 QA 报告写成 XLSX 验收问题清单，返回最终路径。"""

    workbook = Workbook()

    issues_sheet = workbook.active
    issues_sheet.title = "问题清单"
    headers = [
        "页码", "问题类型", "严重度", "描述", "检测器",
        "源区域", "目标区域", "X", "Y", "宽", "高", "人工判定", "备注",
    ]
    issues_sheet.append(headers)
    for cell in issues_sheet[1]:
        cell.font = Font(bold=True)

    for page in report.pages:
        for issue in page.issues:
            issues_sheet.append([
                issue.page,
                issue.type.value,
                issue.severity.value,
                issue.description,
                issue.detector,
                issue.source_region,
                issue.target_region,
                round(issue.bbox.x, 1) if issue.bbox else None,
                round(issue.bbox.y, 1) if issue.bbox else None,
                round(issue.bbox.width, 1) if issue.bbox else None,
                round(issue.bbox.height, 1) if issue.bbox else None,
                None,  # 人工判定（confirmed/false_positive/ignored）
                None,  # 备注
            ])
            fill = _SEVERITY_FILL.get(issue.severity.value)
            if fill:
                row = issues_sheet.max_row
                for column in range(1, len(headers) + 1):
                    issues_sheet.cell(row=row, column=column).fill = fill

    summary_sheet = workbook.create_sheet("文档摘要")
    normalized = report.metadata.get("normalized_from")
    summary_rows = [
        ("文档状态", report.status.value),
        ("文档分数", round(report.document_score, 2)),
        ("页面总数", report.summary.pages),
        ("通过页面", report.summary.passed_pages),
        ("复核页面", report.summary.review_pages),
        ("失败页面", report.summary.failed_pages),
        ("规则配置", report.rule_profile_reference),
        ("问题总数", sum(report.summary.issue_counts.values())),
        ("源文档", report.source_document_id[:16]),
        ("目标文档", report.target_document_id[:16]),
    ]
    if normalized:
        summary_rows.append(("归一化来源", str(normalized)))
    for row in summary_rows:
        summary_sheet.append(row)
    summary_sheet.column_dimensions["A"].width = 14

    output_path = output_path.expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path
